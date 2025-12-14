"""
Output Handler - Writes extracted data to Excel files and CSV files.
"""

import os
import csv
from datetime import datetime
from openpyxl import Workbook, load_workbook

from config import OUTPUT_FILE, get_field_names


def write_to_excel(filename: str, extracted_data: dict, output_path: str = None, source_path: str = None) -> str:
    """
    Write extracted data to an Excel file.
    
    Args:
        filename: Original filename that was processed
        extracted_data: Dictionary of field names to extracted values
        output_path: Optional custom output path (defaults to OUTPUT_FILE from config)
        source_path: Full path to the source file
        
    Returns:
        Path to the output Excel file
    """
    from config import PROGRAMMATIC_FIELDS
    
    output_path = output_path or OUTPUT_FILE
    
    # Check if file exists
    if os.path.exists(output_path):
        workbook = load_workbook(output_path)
        sheet = workbook.active
    else:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Extracted Data"
        
        # Write header row with programmatic fields first, then AI fields
        headers = PROGRAMMATIC_FIELDS + get_field_names()
        for col, header in enumerate(headers, start=1):
            sheet.cell(row=1, column=col, value=header)
    
    # Find next empty row
    next_row = sheet.max_row + 1
    
    # Extract file metadata for programmatic fields
    file_name_no_ext = os.path.splitext(filename)[0]
    file_ext = os.path.splitext(filename)[1]
    
    # Write programmatic fields
    col = 1
    sheet.cell(row=next_row, column=col, value=source_path or filename)  # SourcePathFilenameWksName
    col += 1
    sheet.cell(row=next_row, column=col, value=filename)  # FileName
    col += 1
    sheet.cell(row=next_row, column=col, value=file_name_no_ext)  # FileName wo extension
    col += 1
    sheet.cell(row=next_row, column=col, value=file_ext)  # File Extension
    col += 1
    sheet.cell(row=next_row, column=col, value="")  # Comment (empty for now)
    col += 1
    
    # Write AI-extracted fields
    field_names = get_field_names()
    
    for field_name in field_names:
        value = extracted_data.get(field_name, "Not found")
        
        # Convert lists/arrays to comma-separated strings for Excel
        if isinstance(value, list):
            value = ", ".join(str(item) for item in value)
        
        sheet.cell(row=next_row, column=col, value=value)
        col += 1
    
    # Auto-adjust column widths
    for column in sheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
        sheet.column_dimensions[column_letter].width = adjusted_width
    
    workbook.save(output_path)
    workbook.close()
    
    return output_path


def get_extraction_summary(extracted_data: dict) -> str:
    """
    Generate a summary of extracted data for console output.
    
    Args:
        extracted_data: Dictionary of field names to extracted values
        
    Returns:
        Formatted string summary
    """
    lines = ["Extracted Data:", "-" * 40]
    
    for field, value in extracted_data.items():
        lines.append(f"  {field}: {value}")
    
    lines.append("-" * 40)
    
    return "\n".join(lines)


def write_to_csv(filename: str, extracted_data: dict, output_dir: str = "output_csv", source_path: str = None) -> str:
    """
    Write extracted data to a timestamped CSV file.
    
    Args:
        filename: Original filename that was processed
        extracted_data: Dictionary of field names to extracted values
        output_dir: Directory to save CSV files (created if doesn't exist)
        source_path: Full path to the source file
        
    Returns:
        Path to the created CSV file
    """
    from config import PROGRAMMATIC_FIELDS
    
    # Create output directory if it doesn't exist
    os.makedirs(output_dir, exist_ok=True)
    
    # Generate timestamped filename
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    base_name = os.path.splitext(os.path.basename(filename))[0]
    csv_filename = f"{base_name}_{timestamp}.csv"
    csv_path = os.path.join(output_dir, csv_filename)
    
    # Extract file metadata
    file_name_no_ext = os.path.splitext(filename)[0]
    file_ext = os.path.splitext(filename)[1]
    
    # Prepare data row
    field_names = get_field_names()
    
    # Create CSV with headers and data
    with open(csv_path, 'w', newline='', encoding='utf-8') as csvfile:
        writer = csv.writer(csvfile)
        
        # Write headers: programmatic + AI fields
        headers = PROGRAMMATIC_FIELDS + field_names
        writer.writerow(headers)
        
        # Write data row
        row = [
            source_path or filename,  # SourcePathFilenameWksName
            filename,  # FileName
            file_name_no_ext,  # FileName wo extension
            file_ext,  # File Extension
            "",  # Comment (empty)
        ]
        
        # Add AI-extracted fields
        for field_name in field_names:
            value = extracted_data.get(field_name, "Not found")
            
            # Convert lists to comma-separated strings
            if isinstance(value, list):
                value = ", ".join(str(item) for item in value)
            
            row.append(value)
        
        writer.writerow(row)
    
    return csv_path

