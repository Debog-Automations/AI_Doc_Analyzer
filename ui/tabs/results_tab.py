"""
Results Tab - Display extracted data and export options

Shows data from the database sorted by most recently processed,
with visual distinction between AI-extracted and metadata columns.
"""

from PyQt5.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QGroupBox, QLabel,
    QPushButton, QTableWidget, QTableWidgetItem, QHeaderView,
    QFileDialog, QMessageBox, QComboBox, QLineEdit, QAbstractItemView,
    QRadioButton, QButtonGroup, QSpinBox, QFrame
)
from PyQt5.QtCore import pyqtSignal, Qt, QTimer
from PyQt5.QtGui import QColor, QBrush, QFont
from typing import List, Dict, Any, Optional, Set
import os
import json
from datetime import datetime

# ============================================================================
# COLUMN CLASSIFICATION
# ============================================================================

# AI-extracted columns (from OpenAI via custom questions)
# These get a blue-tinted header background
AI_COLUMNS: Set[str] = {
    "Title", "Type", "AI Summary",
    "As Of Dates", "Effective Date", "Executed Date", "Expiration Date",
    "Currency", "Broker Name", "Carrier Name", "MGA Name", "Intermediary Name",
    "Ceded Percent", "Commission Rates",
    "GWP Actual", "GWP Estimated", "NWP Actual", "NWP Estimated",
    "Lines of Business", "Parties", "Signers", "Sections/Chapters",
    "Countries", "States", "Key Entities", "Table Names", "All Values"
}

# Metadata columns (programmatically extracted)
# These get a gray-tinted header background
METADATA_COLUMNS: Set[str] = {
    "FileName", "file_path", "source_path", "File Extension", "file_extension",
    "Document Size", "Document Size Formatted", "file_size",
    "File Hash", "file_hash", "Has Been Processed",
    "Adobe Document ID", "Docusign Document ID", "E-signature ID",
    "Document Title", "Author", "Subject", "Creator", "Producer",
    "Creation Date", "Modification Date", "Page Count",
    "Comment", "MIME Type", "File Created At", "File Modified At",
    "source_type", "processed_at", "id", "error_message"
}

# Status column - special handling
STATUS_COLUMN = "Status"

# Header colors
AI_HEADER_COLOR = QColor("#E3F2FD")       # Light blue for AI columns
METADATA_HEADER_COLOR = QColor("#F5F5F5")  # Light gray for metadata columns
STATUS_HEADER_COLOR = QColor("#E8F5E9")    # Light green for status


def get_column_category(column_name: str) -> str:
    """
    Determine the category of a column.
    
    Returns:
        'ai' for AI-extracted columns
        'metadata' for programmatic metadata columns
        'status' for status column
        'unknown' for unclassified columns (treated as AI)
    """
    if column_name == STATUS_COLUMN:
        return 'status'
    elif column_name in AI_COLUMNS:
        return 'ai'
    elif column_name in METADATA_COLUMNS:
        return 'metadata'
    else:
        # Unknown columns are assumed to be AI-extracted (custom questions can add new ones)
        return 'ai'


class ResultsTab(QWidget):
    """Tab for displaying and exporting extraction results from database."""
    
    export_requested = pyqtSignal(str)  # Export file path
    
    def __init__(self):
        super().__init__()
        self.results_data: List[Dict[str, Any]] = []
        self.session_hashes: Set[str] = set()  # Track hashes from current session
        
        # Pagination state
        self.current_page = 1
        self.records_per_page = 50
        self.total_records = 0
        
        # Database service (lazy loaded)
        self._dedup_service = None
        
        # View mode
        self.show_all_records = True  # True = all DB records, False = current session only
        
        self._init_ui()
        
        # Load from database after UI is initialized
        QTimer.singleShot(100, self._initial_load)
    
    @property
    def dedup_service(self):
        """Lazy-load the dedup service."""
        if self._dedup_service is None:
            try:
                from services.dedup import create_dedup_service
                # Try to load config
                config = {}
                if os.path.exists("app_config.json"):
                    with open("app_config.json", "r") as f:
                        config = json.load(f)
                self._dedup_service = create_dedup_service(config)
            except Exception as e:
                print(f"Failed to initialize dedup service: {e}")
        return self._dedup_service
    
    def _init_ui(self):
        """Initialize the results UI."""
        layout = QVBoxLayout(self)
        layout.setSpacing(10)
        
        # === Top Controls Section ===
        top_controls = QHBoxLayout()
        
        # Summary label
        self.summary_label = QLabel("Loading...")
        self.summary_label.setStyleSheet("font-size: 14px; font-weight: bold;")
        top_controls.addWidget(self.summary_label)
        
        top_controls.addStretch()
        
        # View mode toggle
        view_group = QButtonGroup(self)
        
        self.all_records_radio = QRadioButton("All Records")
        self.all_records_radio.setChecked(True)
        self.all_records_radio.toggled.connect(self._on_view_mode_changed)
        view_group.addButton(self.all_records_radio)
        top_controls.addWidget(self.all_records_radio)
        
        self.session_only_radio = QRadioButton("Current Session")
        self.session_only_radio.setToolTip("Show only documents processed in this session")
        view_group.addButton(self.session_only_radio)
        top_controls.addWidget(self.session_only_radio)
        
        # Separator
        sep = QFrame()
        sep.setFrameShape(QFrame.VLine)
        sep.setStyleSheet("color: #ccc;")
        top_controls.addWidget(sep)
        
        # Refresh button
        self.refresh_btn = QPushButton("Refresh")
        self.refresh_btn.setMaximumWidth(80)
        self.refresh_btn.clicked.connect(self.load_from_database)
        top_controls.addWidget(self.refresh_btn)
        
        # Filter
        top_controls.addWidget(QLabel("Filter:"))
        self.filter_input = QLineEdit()
        self.filter_input.setPlaceholderText("Type to filter...")
        self.filter_input.setMaximumWidth(200)
        self.filter_input.textChanged.connect(self._apply_filter)
        top_controls.addWidget(self.filter_input)
        
        layout.addLayout(top_controls)
        
        # === Legend for column colors ===
        legend_layout = QHBoxLayout()
        legend_layout.setSpacing(20)
        
        # AI columns legend
        ai_legend = QLabel("■ AI Extracted")
        ai_legend.setStyleSheet(f"color: #1565C0; font-size: 11px;")
        legend_layout.addWidget(ai_legend)
        
        # Metadata columns legend
        meta_legend = QLabel("■ Metadata")
        meta_legend.setStyleSheet(f"color: #616161; font-size: 11px;")
        legend_layout.addWidget(meta_legend)
        
        legend_layout.addStretch()
        layout.addLayout(legend_layout)
        
        # === Results Table ===
        self.table = QTableWidget()
        self.table.setAlternatingRowColors(True)
        self.table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)
        self.table.horizontalHeader().setStretchLastSection(True)
        self.table.verticalHeader().setVisible(False)
        self.table.setSortingEnabled(False)  # We handle sorting via database
        self.table.setStyleSheet("""
            QTableWidget {
                gridline-color: #d0d0d0;
            }
            QTableWidget::item {
                padding: 5px;
            }
            QTableWidget::item:selected {
                background-color: #0078d4;
                color: white;
            }
            QHeaderView::section {
                padding: 6px;
                border: 1px solid #d0d0d0;
                font-weight: bold;
            }
        """)
        layout.addWidget(self.table, 1)
        
        # === Pagination Controls ===
        pagination_layout = QHBoxLayout()
        pagination_layout.setSpacing(10)
        
        # First/Previous buttons
        self.first_btn = QPushButton("<<")
        self.first_btn.setMaximumWidth(40)
        self.first_btn.setToolTip("First page")
        self.first_btn.clicked.connect(self._go_to_first_page)
        pagination_layout.addWidget(self.first_btn)
        
        self.prev_btn = QPushButton("<")
        self.prev_btn.setMaximumWidth(40)
        self.prev_btn.setToolTip("Previous page")
        self.prev_btn.clicked.connect(self._go_to_prev_page)
        pagination_layout.addWidget(self.prev_btn)
        
        # Page info label
        self.page_info_label = QLabel("Page 1 of 1")
        self.page_info_label.setAlignment(Qt.AlignCenter)
        self.page_info_label.setMinimumWidth(150)
        pagination_layout.addWidget(self.page_info_label)
        
        # Next/Last buttons
        self.next_btn = QPushButton(">")
        self.next_btn.setMaximumWidth(40)
        self.next_btn.setToolTip("Next page")
        self.next_btn.clicked.connect(self._go_to_next_page)
        pagination_layout.addWidget(self.next_btn)
        
        self.last_btn = QPushButton(">>")
        self.last_btn.setMaximumWidth(40)
        self.last_btn.setToolTip("Last page")
        self.last_btn.clicked.connect(self._go_to_last_page)
        pagination_layout.addWidget(self.last_btn)
        
        pagination_layout.addStretch()
        
        # Records per page
        pagination_layout.addWidget(QLabel("Records per page:"))
        self.per_page_combo = QComboBox()
        self.per_page_combo.addItems(["25", "50", "100", "200"])
        self.per_page_combo.setCurrentText("50")
        self.per_page_combo.currentTextChanged.connect(self._on_per_page_changed)
        pagination_layout.addWidget(self.per_page_combo)
        
        # Records range label
        self.records_range_label = QLabel("")
        self.records_range_label.setStyleSheet("color: #666;")
        pagination_layout.addWidget(self.records_range_label)
        
        layout.addLayout(pagination_layout)
        
        # === Export Section ===
        export_group = QGroupBox("Export Options")
        export_layout = QHBoxLayout()
        
        export_layout.addWidget(QLabel("Format:"))
        self.export_format_combo = QComboBox()
        self.export_format_combo.addItems(["Excel (.xlsx)", "CSV (.csv)"])
        self.export_format_combo.setMinimumWidth(150)
        export_layout.addWidget(self.export_format_combo)
        
        export_layout.addStretch()
        
        self.export_btn = QPushButton("Export Results")
        self.export_btn.setMinimumWidth(150)
        self.export_btn.clicked.connect(self._export_results)
        self.export_btn.setEnabled(False)
        export_layout.addWidget(self.export_btn)
        
        self.export_all_btn = QPushButton("Export to Master File")
        self.export_all_btn.setMinimumWidth(150)
        self.export_all_btn.clicked.connect(self._export_to_master)
        self.export_all_btn.setEnabled(False)
        self.export_all_btn.setToolTip("Append results to the master Excel file")
        export_layout.addWidget(self.export_all_btn)
        
        export_group.setLayout(export_layout)
        layout.addWidget(export_group)
    
    def _initial_load(self):
        """Initial load from database."""
        self.load_from_database()
    
    def _on_view_mode_changed(self, checked: bool):
        """Handle view mode toggle."""
        if checked:  # all_records_radio was toggled on
            self.show_all_records = True
        else:
            self.show_all_records = False
        
        self.current_page = 1
        self.load_from_database()
    
    def _on_per_page_changed(self, value: str):
        """Handle records per page change."""
        try:
            self.records_per_page = int(value)
            self.current_page = 1
            self.load_from_database()
        except ValueError:
            pass
    
    def load_from_database(self):
        """Load records from the database."""
        if self.dedup_service is None:
            self.summary_label.setText("Database connection unavailable")
            return
        
        try:
            # Determine which hashes to filter by
            session_hashes = list(self.session_hashes) if not self.show_all_records else None
            
            # Get paginated data
            records, total = self.dedup_service.get_documents_paginated(
                page=self.current_page,
                per_page=self.records_per_page,
                session_hashes=session_hashes
            )
            
            self.results_data = records
            self.total_records = total
            
            # Refresh display
            self._refresh_table()
            self._update_summary()
            self._update_pagination_controls()
            
            # Enable export if we have data
            has_data = len(self.results_data) > 0
            self.export_btn.setEnabled(has_data)
            self.export_all_btn.setEnabled(has_data)
            
        except Exception as e:
            self.summary_label.setText(f"Error loading data: {str(e)}")
            print(f"Database load error: {e}")
    
    def _get_ordered_columns(self, all_keys: Set[str]) -> List[str]:
        """
        Get columns in the correct order:
        1. Status, FileName (always first)
        2. AI columns grouped together
        3. Metadata columns grouped together
        """
        # Remove internal/hidden keys
        display_keys = set(k for k in all_keys if not k.startswith('_'))
        
        # Priority columns (always first)
        priority_cols = ["Status", "FileName"]
        
        # Separate into categories
        ai_cols = []
        metadata_cols = []
        
        for key in display_keys:
            if key in priority_cols:
                continue
            
            category = get_column_category(key)
            if category in ('ai', 'unknown'):
                ai_cols.append(key)
            else:
                metadata_cols.append(key)
        
        # Sort within categories
        ai_cols.sort()
        metadata_cols.sort()
        
        # Build final order
        ordered = []
        for col in priority_cols:
            if col in display_keys:
                ordered.append(col)
        
        ordered.extend(ai_cols)
        ordered.extend(metadata_cols)
        
        return ordered
    
    def _refresh_table(self):
        """Refresh the table with current data."""
        self.table.setRowCount(0)
        
        if not self.results_data:
            self.table.setColumnCount(1)
            self.table.setHorizontalHeaderLabels(["No data"])
            return
        
        # Get all unique keys from results
        all_keys = set()
        for result in self.results_data:
            all_keys.update(result.keys())
        
        # Get ordered columns
        ordered_cols = self._get_ordered_columns(all_keys)
        
        # Setup table
        self.table.setRowCount(len(self.results_data))
        self.table.setColumnCount(len(ordered_cols))
        self.table.setHorizontalHeaderLabels(ordered_cols)
        
        # Apply header colors
        header = self.table.horizontalHeader()
        for col_idx, col_name in enumerate(ordered_cols):
            header_item = self.table.horizontalHeaderItem(col_idx)
            if header_item:
                category = get_column_category(col_name)
                if category == 'ai':
                    header_item.setBackground(AI_HEADER_COLOR)
                    header_item.setForeground(QBrush(QColor("#1565C0")))
                elif category == 'metadata':
                    header_item.setBackground(METADATA_HEADER_COLOR)
                    header_item.setForeground(QBrush(QColor("#424242")))
                elif category == 'status':
                    header_item.setBackground(STATUS_HEADER_COLOR)
                    header_item.setForeground(QBrush(QColor("#2E7D32")))
        
        # Populate data
        for row, result in enumerate(self.results_data):
            for col, key in enumerate(ordered_cols):
                value = result.get(key, "")
                
                # Handle different value types
                if isinstance(value, list):
                    value = ", ".join(str(v) for v in value)
                elif isinstance(value, dict):
                    if "error" in value:
                        value = f"ERROR: {value['error']}"
                    else:
                        value = str(value)
                else:
                    value = str(value) if value is not None else ""
                
                item = QTableWidgetItem(value)
                
                # Color code status column
                if key == "Status":
                    status_lower = value.lower()
                    if status_lower == "success" or status_lower == "completed":
                        item.setBackground(QColor("#d4edda"))
                    elif "error" in status_lower or "failed" in status_lower:
                        item.setBackground(QColor("#f8d7da"))
                    elif "skipped" in status_lower:
                        item.setBackground(QColor("#fff3cd"))
                
                # Color code errors in any column
                if "error" in str(result.get("error_message", "")).lower() or \
                   "ERROR" in value:
                    item.setForeground(QColor("#dc3545"))
                
                self.table.setItem(row, col, item)
        
        # Auto-resize columns (with max width limit)
        self.table.resizeColumnsToContents()
        for col in range(self.table.columnCount()):
            if self.table.columnWidth(col) > 300:
                self.table.setColumnWidth(col, 300)
    
    def _update_summary(self):
        """Update the summary label."""
        total = self.total_records
        
        if self.results_data:
            success = sum(1 for r in self.results_data 
                         if r.get("Status", "").lower() in ("success", "completed"))
            errors = sum(1 for r in self.results_data 
                        if "error" in r.get("Status", "").lower() or "failed" in r.get("Status", "").lower())
            skipped = sum(1 for r in self.results_data 
                         if "skipped" in r.get("Status", "").lower())
            
            parts = [f"Total: {total}"]
            if success:
                parts.append(f"{success} successful")
            if errors:
                parts.append(f"{errors} errors")
            if skipped:
                parts.append(f"{skipped} skipped")
            
            self.summary_label.setText(" | ".join(parts))
        else:
            self.summary_label.setText("No results in database")
    
    def _update_pagination_controls(self):
        """Update pagination button states and labels."""
        total_pages = max(1, (self.total_records + self.records_per_page - 1) // self.records_per_page)
        
        # Update page info
        self.page_info_label.setText(f"Page {self.current_page} of {total_pages}")
        
        # Calculate record range
        start_record = (self.current_page - 1) * self.records_per_page + 1
        end_record = min(self.current_page * self.records_per_page, self.total_records)
        
        if self.total_records > 0:
            self.records_range_label.setText(f"(showing {start_record}-{end_record} of {self.total_records})")
        else:
            self.records_range_label.setText("")
        
        # Enable/disable buttons
        self.first_btn.setEnabled(self.current_page > 1)
        self.prev_btn.setEnabled(self.current_page > 1)
        self.next_btn.setEnabled(self.current_page < total_pages)
        self.last_btn.setEnabled(self.current_page < total_pages)
    
    def _go_to_first_page(self):
        """Go to the first page."""
        if self.current_page != 1:
            self.current_page = 1
            self.load_from_database()
    
    def _go_to_prev_page(self):
        """Go to the previous page."""
        if self.current_page > 1:
            self.current_page -= 1
            self.load_from_database()
    
    def _go_to_next_page(self):
        """Go to the next page."""
        total_pages = max(1, (self.total_records + self.records_per_page - 1) // self.records_per_page)
        if self.current_page < total_pages:
            self.current_page += 1
            self.load_from_database()
    
    def _go_to_last_page(self):
        """Go to the last page."""
        total_pages = max(1, (self.total_records + self.records_per_page - 1) // self.records_per_page)
        if self.current_page != total_pages:
            self.current_page = total_pages
            self.load_from_database()
    
    def _apply_filter(self, text: str):
        """Apply filter to table rows."""
        text = text.lower()
        
        for row in range(self.table.rowCount()):
            show = not text  # Show all if filter is empty
            if text:
                for col in range(self.table.columnCount()):
                    item = self.table.item(row, col)
                    if item and text in item.text().lower():
                        show = True
                        break
            self.table.setRowHidden(row, not show)
    
    def add_result(self, file_path: str, data: Dict[str, Any]):
        """Add a single result to the table (and track for session filter)."""
        # Track session hash
        file_hash = data.get("file_hash") or data.get("File Hash", "")
        if file_hash:
            self.session_hashes.add(file_hash)
        
        # Reload from database to show updated data
        self.load_from_database()
    
    def add_results(self, results: List[tuple]):
        """Add multiple results. Each tuple is (file_path, data_dict)."""
        for file_path, data in results:
            file_hash = data.get("file_hash") or data.get("File Hash", "")
            if file_hash:
                self.session_hashes.add(file_hash)
        
        # Reload from database
        self.load_from_database()
    
    def _export_results(self):
        """Export results to a new file."""
        if not self.results_data:
            QMessageBox.warning(self, "No Data", "No results to export.")
            return
        
        # Determine file format
        is_excel = self.export_format_combo.currentIndex() == 0
        ext = ".xlsx" if is_excel else ".csv"
        filter_str = "Excel Files (*.xlsx)" if is_excel else "CSV Files (*.csv)"
        
        # Generate default filename
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        default_name = f"extraction_results_{timestamp}{ext}"
        
        file_path, _ = QFileDialog.getSaveFileName(
            self, "Export Results", default_name, filter_str
        )
        
        if file_path:
            try:
                self._write_export(file_path, is_excel)
                QMessageBox.information(
                    self, "Export Complete", 
                    f"Results exported to:\n{file_path}"
                )
            except Exception as e:
                QMessageBox.critical(
                    self, "Export Error",
                    f"Failed to export: {str(e)}"
                )
    
    def _export_to_master(self):
        """Export results to the master Excel file."""
        if not self.results_data:
            QMessageBox.warning(self, "No Data", "No results to export.")
            return
        
        # Use default master file path
        master_path = "extracted_data.xlsx"
        
        try:
            self._write_export(master_path, is_excel=True, append=True)
            QMessageBox.information(
                self, "Export Complete",
                f"Results appended to:\n{master_path}"
            )
        except Exception as e:
            QMessageBox.critical(
                self, "Export Error",
                f"Failed to export: {str(e)}"
            )
    
    def _write_export(self, file_path: str, is_excel: bool, append: bool = False):
        """Write results to file with formatting."""
        
        # Clean up internal fields from export
        export_data = []
        for result in self.results_data:
            cleaned = {k: v for k, v in result.items() if not k.startswith('_')}
            export_data.append(cleaned)
        
        if is_excel:
            self._write_excel(file_path, export_data, append)
        else:
            self._write_csv(file_path, export_data)
    
    def _write_excel(self, file_path: str, data: list, append: bool = False):
        """Write data to Excel with formatting."""
        from openpyxl import Workbook, load_workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        import pandas as pd
        
        df = pd.DataFrame(data)
        
        if append and os.path.exists(file_path):
            # Append to existing
            existing_df = pd.read_excel(file_path)
            df = pd.concat([existing_df, df], ignore_index=True)
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Extracted Data"
        
        # Write headers with color coding
        headers = list(df.columns)
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            
            # Color based on category
            category = get_column_category(header)
            if category == 'ai':
                cell.fill = PatternFill(start_color="1565C0", end_color="1565C0", fill_type="solid")
            elif category == 'metadata':
                cell.fill = PatternFill(start_color="616161", end_color="616161", fill_type="solid")
            else:
                cell.fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
        
        # Write data
        for row_idx, row_data in enumerate(df.values, 2):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=str(value) if value else "")
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                
                # Color code status column
                if headers[col_idx - 1] == "Status":
                    if "Success" in str(value) or "Completed" in str(value):
                        cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    elif "Error" in str(value) or "Failed" in str(value):
                        cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                    elif "Skipped" in str(value):
                        cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        
        # Auto-adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width
        
        # Freeze header row
        ws.freeze_panes = "A2"
        
        wb.save(file_path)
    
    def _write_csv(self, file_path: str, data: list):
        """Write data to CSV."""
        import pandas as pd
        df = pd.DataFrame(data)
        df.to_csv(file_path, index=False)
    
    def clear_results(self):
        """Clear session tracking and refresh from database."""
        self.session_hashes.clear()
        self.current_page = 1
        self.load_from_database()
    
    def get_results(self) -> List[Dict[str, Any]]:
        """Get all results data currently displayed."""
        return self.results_data.copy()
