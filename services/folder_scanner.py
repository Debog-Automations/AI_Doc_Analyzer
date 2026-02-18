"""
Folder Scanner Service - Scan folders for new documents to process

Recursively scans configured folders, identifies new files by comparing
against the dedup database, and returns files ready for processing.
Supports both local folders and Box cloud storage.
"""

import os
import json
from datetime import datetime
from typing import List, Dict, Any, Optional, Callable
from dataclasses import dataclass, field

from connectors.local_connector import LocalConnector
from connectors.base_connector import FileInfo
from .dedup import DedupService
from .hasher import FileHasher


# Supported file extensions for document processing
SUPPORTED_EXTENSIONS = ['.pdf', '.xlsx', '.xlsm', '.xls']


@dataclass
class ScanFileInfo:
    """
    Information about a file to be processed.
    Supports both local files and cloud files (Box).
    """
    name: str
    path: str  # Local path or cloud path
    source_type: str  # 'local' or 'box'
    file_hash: str  # SHA256 for local, SHA1 for Box
    file_id: Optional[str] = None  # Box file ID (for downloading)
    size: int = 0
    
    def to_dict(self) -> Dict[str, Any]:
        return {
            'name': self.name,
            'path': self.path,
            'source_type': self.source_type,
            'file_hash': self.file_hash,
            'file_id': self.file_id,
            'size': self.size
        }


@dataclass
class ScanResult:
    """Result of a folder scan operation."""
    total_files_found: int
    new_files: List[ScanFileInfo]  # Files to process
    skipped_files: List[str]  # Already processed (in dedup database)
    error_files: List[Dict[str, str]]  # Files that couldn't be checked
    scan_time: datetime
    folders_scanned: List[str]
    source_type: str = "local"  # 'local' or 'box'
    
    def to_dict(self) -> Dict[str, Any]:
        """Convert to dictionary."""
        return {
            'total_files_found': self.total_files_found,
            'new_files_count': len(self.new_files),
            'skipped_files_count': len(self.skipped_files),
            'error_files_count': len(self.error_files),
            'scan_time': self.scan_time.isoformat(),
            'folders_scanned': self.folders_scanned,
            'source_type': self.source_type
        }
    
    def get_new_file_paths(self) -> List[str]:
        """Get list of file paths (for backward compatibility)."""
        return [f.path for f in self.new_files]


class FolderScanner:
    """
    Service for scanning folders and identifying new documents to process.
    
    Supports:
    - Local folders via LocalConnector
    - Box cloud storage via BoxConnector
    
    Uses DedupService for checking which files have already been processed.
    """
    
    def __init__(
        self,
        dedup_service: Optional[DedupService] = None,
        extensions: Optional[List[str]] = None
    ):
        """
        Initialize the folder scanner.
        
        Args:
            dedup_service: Optional DedupService instance. Creates one if not provided.
            extensions: List of file extensions to include. Defaults to SUPPORTED_EXTENSIONS.
        """
        self.dedup_service = dedup_service or DedupService()
        self.extensions = extensions or SUPPORTED_EXTENSIONS
        self.local_connector = LocalConnector()
        self._box_connector = None
    
    def scan_folders(
        self,
        folder_paths: List[str],
        progress_callback: Optional[Callable[[str, int, int], None]] = None
    ) -> ScanResult:
        """
        Scan multiple local folders for new documents.
        
        Args:
            folder_paths: List of folder paths to scan
            progress_callback: Optional callback(message, current, total) for progress updates
            
        Returns:
            ScanResult with lists of new and skipped files
        """
        all_files: List[FileInfo] = []
        valid_folders: List[str] = []
        
        # Collect all files from all folders
        for folder_path in folder_paths:
            if not os.path.exists(folder_path):
                if progress_callback:
                    progress_callback(f"Folder not found: {folder_path}", 0, 0)
                continue
            
            if not os.path.isdir(folder_path):
                if progress_callback:
                    progress_callback(f"Not a directory: {folder_path}", 0, 0)
                continue
            
            valid_folders.append(folder_path)
            
            # Use LocalConnector to list files recursively
            files = self.local_connector.list_files(
                path=folder_path,
                extensions=self.extensions,
                recursive=True
            )
            all_files.extend(files)
        
        if progress_callback:
            progress_callback(f"Found {len(all_files)} files in {len(valid_folders)} folders", 0, len(all_files))
        
        # Check each file against dedup database
        new_files: List[ScanFileInfo] = []
        skipped_files: List[str] = []
        error_files: List[Dict[str, str]] = []
        
        for idx, file_info in enumerate(all_files):
            file_path = file_info.path
            
            if progress_callback:
                progress_callback(f"Checking: {file_info.name}", idx + 1, len(all_files))
            
            try:
                # Compute hash and check if already processed
                file_hash = FileHasher.hash_file(file_path)
                
                if self.dedup_service.hash_exists(file_hash):
                    skipped_files.append(file_path)
                else:
                    new_files.append(ScanFileInfo(
                        name=file_info.name,
                        path=file_path,
                        source_type='local',
                        file_hash=file_hash,
                        file_id=None,
                        size=file_info.size
                    ))
                    
            except Exception as e:
                error_files.append({
                    'path': file_path,
                    'error': str(e)
                })
        
        return ScanResult(
            total_files_found=len(all_files),
            new_files=new_files,
            skipped_files=skipped_files,
            error_files=error_files,
            scan_time=datetime.now(),
            folders_scanned=valid_folders,
            source_type='local'
        )
    
    def scan_box_folders(
        self,
        folder_paths: List[str],
        box_connector,
        progress_callback: Optional[Callable[[str, int, int], None]] = None
    ) -> ScanResult:
        """
        Scan Box folders for new documents.
        
        Uses Box's built-in SHA1 hash for dedup checking (no download required).
        
        Args:
            folder_paths: List of Box folder paths to scan (e.g., "/Documents/Contracts")
            box_connector: Connected BoxConnector instance
            progress_callback: Optional callback(message, current, total) for progress updates
            
        Returns:
            ScanResult with lists of new files (including Box file IDs for downloading)
        """
        all_files: List[FileInfo] = []
        valid_folders: List[str] = []
        
        # Collect all files from all folders
        for folder_path in folder_paths:
            if progress_callback:
                progress_callback(f"Scanning Box folder: {folder_path}", 0, 0)
            
            try:
                # Use BoxConnector to list files recursively
                files = box_connector.list_files(
                    path=folder_path,
                    extensions=self.extensions,
                    recursive=True
                )
                all_files.extend(files)
                valid_folders.append(folder_path)
            except Exception as e:
                if progress_callback:
                    progress_callback(f"Error scanning {folder_path}: {str(e)}", 0, 0)
        
        if progress_callback:
            progress_callback(f"Found {len(all_files)} files in {len(valid_folders)} Box folders", 0, len(all_files))
        
        # Check each file against dedup database using Box's SHA1 hash
        new_files: List[ScanFileInfo] = []
        skipped_files: List[str] = []
        error_files: List[Dict[str, str]] = []
        
        for idx, file_info in enumerate(all_files):
            if progress_callback:
                progress_callback(f"Checking: {file_info.name}", idx + 1, len(all_files))
            
            try:
                # Use Box's built-in SHA1 hash (no download needed!)
                file_hash = file_info.hash
                
                if not file_hash:
                    # If hash not available, we can't check - treat as new
                    new_files.append(ScanFileInfo(
                        name=file_info.name,
                        path=file_info.path,
                        source_type='box',
                        file_hash='',
                        file_id=file_info.id,
                        size=file_info.size
                    ))
                elif self.dedup_service.hash_exists(file_hash):
                    skipped_files.append(file_info.path)
                else:
                    new_files.append(ScanFileInfo(
                        name=file_info.name,
                        path=file_info.path,
                        source_type='box',
                        file_hash=file_hash,
                        file_id=file_info.id,
                        size=file_info.size
                    ))
                    
            except Exception as e:
                error_files.append({
                    'path': file_info.path,
                    'error': str(e)
                })
        
        return ScanResult(
            total_files_found=len(all_files),
            new_files=new_files,
            skipped_files=skipped_files,
            error_files=error_files,
            scan_time=datetime.now(),
            folders_scanned=valid_folders,
            source_type='box'
        )
    
    def scan_single_folder(
        self,
        folder_path: str,
        progress_callback: Optional[Callable[[str, int, int], None]] = None
    ) -> ScanResult:
        """
        Scan a single local folder for new documents.
        
        Args:
            folder_path: Path to folder to scan
            progress_callback: Optional callback for progress updates
            
        Returns:
            ScanResult with lists of new and skipped files
        """
        return self.scan_folders([folder_path], progress_callback)
    
    def get_file_count(self, folder_paths: List[str]) -> int:
        """
        Get total count of supported files in local folders without checking dedup.
        
        Args:
            folder_paths: List of folder paths to count
            
        Returns:
            Total number of supported files
        """
        total = 0
        for folder_path in folder_paths:
            if os.path.exists(folder_path) and os.path.isdir(folder_path):
                files = self.local_connector.list_files(
                    path=folder_path,
                    extensions=self.extensions,
                    recursive=True
                )
                total += len(files)
        return total


def save_result_to_json(
    result: Dict[str, Any],
    file_path: str,
    output_folder: str
) -> str:
    """
    Save extraction result to a JSON file.
    
    Args:
        result: Extraction result dictionary
        file_path: Original file path (used for naming)
        output_folder: Folder to save JSON files
        
    Returns:
        Path to saved JSON file
    """
    os.makedirs(output_folder, exist_ok=True)
    
    base_name = os.path.splitext(os.path.basename(file_path))[0]
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = os.path.join(output_folder, f"{base_name}_{timestamp}.json")
    
    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(result, f, indent=2, ensure_ascii=False, default=str)
    
    return output_path


def get_metadata_columns() -> List[str]:
    """
    Get the ordered list of metadata column names.
    These are extracted programmatically, not via AI.
    """
    return [
        "SourcePathFilenameWksName",
        "FileName",
        "FileName wo extension",
        "File Extension",
        "Document Size",
        "Document Size Formatted",
        "File Hash",
        "Has Been Processed",
        "MIME Type",
        "File Created At",
        "File Modified At",
        "Page Count",
        "Document Title",  # Renamed from "Title" to avoid conflict with AI Title
        "Author",
        "Subject",
        "Creator",
        "Producer",
        "Creation Date",
        "Modification Date",
        "Adobe Document ID",
        "Docusign Document ID",
        "E-signature ID",
        "Comment",
    ]


def get_ai_columns() -> List[str]:
    """
    Get the ordered list of AI column names.
    These are extracted via OpenAI.
    """
    return [
        # Core fields
        "Title",
        "Type",
        "AsOfDt",
        "ExecutedDt",
        "EffDt",
        "ExpDt",
        "Currency",
        # Reinsurance specific
        "Broker Name",
        "Intermediary Name",
        "MGA Name",
        "GWP Actual",
        "GWP Est",
        "NWP Actual",
        "NWP Est",
        "Ceded Percent",
        # MGA/Broker specific
        "Carrier Name",
        "Commission Rates",
        "Territory",
        "Lines of Business",
        # Generic
        "Key Entities",
        "Financial Terms",
        # Parties
        "Party 1 Name",
        "Party 2 Name",
        "Party 3 Name",
        "Party 4 Name",
        "Party 5 Name",
        "Party 6 Name",
        "Party 7 Name",
        "Party 8 Name",
        "Party 9 Name",
        "Party 10 Name",
        # Tables
        "Table 1 Name",
        "Table 2 Name",
        "Table 3 Name",
        "Table 4 Name",
        "Table 5 Name",
        "Table 6 Name",
        "Table 7 Name",
        "Table 8 Name",
        "Table 9 Name",
        "Table 10 Name",
        # Sections
        "Section 1 Name",
        "Section 2 Name",
        "Section 3 Name",
        "Section 4 Name",
        "Section 5 Name",
        "Section 6 Name",
        "Section 7 Name",
        "Section 8 Name",
        "Section 9 Name",
        "Section 10 Name",
        # Summary
        "AI Summarize the Document",
        # Processing metadata
        "Status",
        "_ai_retries",
        "_ai_missing",
    ]


def get_all_columns() -> List[str]:
    """Get all columns in order: metadata first, then AI."""
    return get_metadata_columns() + get_ai_columns()


def normalize_result_for_export(result: Dict[str, Any]) -> Dict[str, Any]:
    """
    Normalize result dictionary to match expected column names.
    Ensures all expected columns exist with default empty values.
    """
    normalized = {}
    all_columns = get_all_columns()
    
    # Copy all values from result
    for key, value in result.items():
        normalized[key] = value
    
    # Ensure all columns exist (with empty string for missing)
    for col in all_columns:
        if col not in normalized:
            normalized[col] = ""
    
    return normalized


def append_result_to_excel(
    result: Dict[str, Any],
    excel_path: str
) -> str:
    """
    Append extraction result to an Excel file with color-coded headers.
    
    Creates the file if it doesn't exist, otherwise appends a new row.
    Headers are color-coded:
    - Blue background: Metadata fields (extracted programmatically)
    - Green background: AI fields (extracted via OpenAI)
    
    Args:
        result: Extraction result dictionary
        excel_path: Path to the Excel file
        
    Returns:
        Path to the Excel file
    """
    import pandas as pd
    from openpyxl import load_workbook, Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
    
    # Ensure output folder exists
    output_dir = os.path.dirname(excel_path)
    if output_dir:
        os.makedirs(output_dir, exist_ok=True)
    
    # Define column order
    metadata_cols = get_metadata_columns()
    ai_cols = get_ai_columns()
    all_columns = metadata_cols + ai_cols
    
    # Normalize the result
    normalized_result = normalize_result_for_export(result)
    
    # Create ordered row data
    row_data = {col: normalized_result.get(col, "") for col in all_columns}
    
    # Define styles
    metadata_header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")  # Blue
    ai_header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")  # Green
    header_font = Font(bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    if os.path.exists(excel_path):
        # Load existing workbook
        try:
            wb = load_workbook(excel_path)
            ws = wb.active
            
            # Append the new row at the end
            next_row = ws.max_row + 1
            for col_idx, col_name in enumerate(all_columns, 1):
                value = row_data.get(col_name, "")
                ws.cell(row=next_row, column=col_idx, value=value)
                ws.cell(row=next_row, column=col_idx).border = thin_border
            
        except Exception:
            # If file is corrupted, start fresh
            wb = Workbook()
            ws = wb.active
            _write_new_excel(ws, all_columns, metadata_cols, row_data,
                           metadata_header_fill, ai_header_fill, header_font,
                           header_alignment, thin_border)
    else:
        # Create new workbook
        wb = Workbook()
        ws = wb.active
        _write_new_excel(ws, all_columns, metadata_cols, row_data,
                        metadata_header_fill, ai_header_fill, header_font,
                        header_alignment, thin_border)
    
    # Save workbook
    wb.save(excel_path)
    
    return excel_path


def _write_new_excel(ws, all_columns, metadata_cols, row_data,
                     metadata_header_fill, ai_header_fill, header_font,
                     header_alignment, thin_border):
    """Helper to write a new Excel file with headers and first data row."""
    
    # Write headers with formatting
    for col_idx, col_name in enumerate(all_columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        
        # Apply color based on whether it's metadata or AI
        if col_name in metadata_cols:
            cell.fill = metadata_header_fill
        else:
            cell.fill = ai_header_fill
        
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Write data row
    for col_idx, col_name in enumerate(all_columns, 1):
        value = row_data.get(col_name, "")
        cell = ws.cell(row=2, column=col_idx, value=value)
        cell.border = thin_border
    
    # Auto-adjust column widths (approximate)
    for col_idx, col_name in enumerate(all_columns, 1):
        # Set a reasonable width based on header length
        width = min(max(len(col_name) + 2, 12), 50)
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = width
    
    # Freeze the header row
    ws.freeze_panes = 'A2'

