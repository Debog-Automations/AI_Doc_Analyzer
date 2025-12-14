"""
Excel Extractor - Reads XLSM/XLSX files and converts to structured text for AI processing.
"""

from openpyxl import load_workbook


def extract_excel_content(excel_path: str, max_rows_per_sheet: int = 1000) -> str:
    """
    Extract content from an Excel file and format as structured text.
    
    Args:
        excel_path: Path to the Excel file (.xlsx or .xlsm)
        max_rows_per_sheet: Maximum rows to read per sheet (to avoid token limits)
        
    Returns:
        Formatted string representation of the Excel content
    """
    workbook = load_workbook(excel_path, read_only=True, data_only=True)
    
    content_parts = []
    
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        
        sheet_content = [f"\n=== Sheet: {sheet_name} ===\n"]
        
        row_count = 0
        for row in sheet.iter_rows(values_only=True):
            if row_count >= max_rows_per_sheet:
                sheet_content.append(f"... (truncated, showing first {max_rows_per_sheet} rows)")
                break
            
            # Convert row to string, handling None values
            row_values = [str(cell) if cell is not None else "" for cell in row]
            
            # Skip completely empty rows
            if not any(row_values):
                continue
            
            # Format as tab-separated for readability
            sheet_content.append(" | ".join(row_values))
            row_count += 1
        
        if row_count > 0:
            content_parts.append("\n".join(sheet_content))
    
    workbook.close()
    
    return "\n".join(content_parts)


def get_sheet_names(excel_path: str) -> list[str]:
    """Get list of sheet names in an Excel file."""
    workbook = load_workbook(excel_path, read_only=True)
    names = workbook.sheetnames
    workbook.close()
    return names

